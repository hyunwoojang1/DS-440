import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Task definitions
# status: "DONE" | "IN PROGRESS" | "PENDING"
# ─────────────────────────────────────────────
tasks = [
    # (id, name, phase, duration, predecessors, status)
    (1,  "Requirements Analysis",                    "Planning",       2,  [],          "DONE"),
    (2,  "System Architecture Design",               "Planning",       2,  [1],         "DONE"),
    (3,  "Functional Decomposition – Level 0",       "Planning",       1,  [2],         "DONE"),
    (4,  "Functional Decomposition – Level 1",       "Planning",       2,  [3],         "DONE"),
    (5,  "Project Setup & Directory Structure",      "Setup",          1,  [4],         "DONE"),
    (6,  "Config Files (weights, series)",           "Setup",          2,  [5],         "DONE"),
    (7,  "Disk Cache (Parquet + TTL)",               "Data Acq.",      2,  [5],         "DONE"),
    (8,  "FRED Fetcher (10 macro series)",           "Data Acq.",      3,  [7],         "DONE"),
    (9,  "WRDS Fetcher (7 Compustat metrics)",       "Data Acq.",      3,  [7],         "DONE"),
    (10, "Technical Fetcher (yfinance, 8 indicators)","Data Acq.",     3,  [7],         "DONE"),
    (11, "Ticker Resolver (yfinance Search)",        "Data Acq.",      1,  [5],         "DONE"),
    (12, "BaseNormalizer & Look-ahead Bias Guard",   "Normalization",  2,  [8,9,10],    "DONE"),
    (13, "MinMax Normalizer",                        "Normalization",  1,  [12],        "DONE"),
    (14, "Z-Score Normalizer",                       "Normalization",  1,  [12],        "DONE"),
    (15, "Percentile Rank Normalizer",               "Normalization",  1,  [12],        "DONE"),
    (16, "Normalization Integration Tests",          "Normalization",  1,  [13,14,15],  "DONE"),
    (17, "MacroScorer",                              "Scoring",        2,  [16,6],      "DONE"),
    (18, "FundamentalScorer (GICS sectors)",         "Scoring",        3,  [16,6],      "DONE"),
    (19, "TechnicalScorer (V-shape RSI)",            "Scoring",        2,  [16,6],      "DONE"),
    (20, "Scorer Integration Tests",                 "Scoring",        1,  [17,18,19],  "DONE"),
    (21, "HorizonResult Dataclass",                  "Aggregation",    1,  [20],        "DONE"),
    (22, "ShortTermHorizon Aggregator",              "Aggregation",    1,  [21],        "DONE"),
    (23, "MidTermHorizon Aggregator",                "Aggregation",    1,  [21],        "DONE"),
    (24, "LongTermHorizon Aggregator",               "Aggregation",    1,  [21],        "DONE"),
    (25, "HTML Dashboard Formatter",                 "Output",         3,  [22,23,24],  "DONE"),
    (26, "JSON / CSV Formatter",                     "Output",         1,  [22,23,24],  "DONE"),
    (27, "Terminal Table (rich library)",            "Output",         1,  [22,23,24],  "DONE"),
    (28, "CLI Entry Point (main.py)",                "Output",         2,  [25,26,27,11],"DONE"),
    (29, "End-to-End Testing",                       "Testing",        2,  [28],        "DONE"),
    (30, "Web Application (Flask UI)",               "Web",            3,  [28],        "DONE"),
    (31, "Final Review & Documentation",             "Testing",        2,  [29,30],     "IN PROGRESS"),
]

# Status styling
STATUS_COLORS = {
    "DONE":        ("✔ DONE",        "1E4620", "D4EDDA"),   # (label, font_hex, bg_hex)
    "IN PROGRESS": ("⏳ IN PROGRESS", "7F4F24", "FFF3CD"),
    "PENDING":     ("○ PENDING",     "4D4D4D", "F2F2F2"),
}

# ─────────────────────────────────────────────
# Compute ES / EF / LS / LF / Slack
# ─────────────────────────────────────────────
task_map = {t[0]: list(t) for t in tasks}
# index: 0=id 1=name 2=phase 3=dur 4=preds 5=status, then append: ES EF LS LF slack critical
for t in task_map.values():
    t += [0, 0, 0, 0, 0, False]  # ES=6, EF=7, LS=8, LF=9, slack=10, critical=11

# Forward pass
for tid, t in task_map.items():
    preds = t[4]
    es = max((task_map[p][7] for p in preds), default=0)
    t[6] = es          # ES
    t[7] = es + t[3]   # EF

total = max(t[7] for t in task_map.values())

# Backward pass — init LF
for t in task_map.values():
    t[9] = total

for tid in sorted(task_map.keys(), key=lambda x: -task_map[x][7]):
    t = task_map[tid]
    t[8] = t[9] - t[3]   # LS
    for p in t[4]:
        task_map[p][9] = min(task_map[p][9], t[8])

for t in task_map.values():
    t[10] = t[8] - t[6]     # slack
    t[11] = (t[10] == 0)    # critical

num_tasks = len(tasks)

# ─────────────────────────────────────────────
# Color palette
# ─────────────────────────────────────────────
PHASE_COLORS = {
    "Planning":      "1F4E79",
    "Setup":         "375623",
    "Data Acq.":     "7B2C2C",
    "Normalization": "4A235A",
    "Scoring":       "1C4587",
    "Aggregation":   "7F4F24",
    "Output":        "0C5E3E",
    "Web":           "1A5276",
    "Testing":       "4D4D00",
}
GANTT_BAR      = "4472C4"
GANTT_CRIT     = "C00000"
GANTT_DONE     = "2E7D32"   # green bar for completed tasks
GANTT_DONE_CRIT= "1B5E20"   # dark green for completed critical tasks
HEADER_FILL    = "2F2F2F"
ALT_ROW        = "F2F2F2"
WHITE          = "FFFFFF"

thin   = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_fill(ws, row, col, hex_color):
    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=hex_color)


def set_cell(ws, row, col, value=None, bold=False, color="000000",
             bg=None, align="left", wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.border = border
    return c


# ══════════════════════════════════════════════
#  GANTT CHART
# ══════════════════════════════════════════════
wb_g = openpyxl.Workbook()
ws = wb_g.active
ws.title = "Gantt Chart"

ws.row_dimensions[1].height = 30

# ── Fixed columns ──────────────────────────────
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 9
ws.column_dimensions["E"].width = 14   # Status column

headers_fixed = ["#", "Task Name", "Phase", "Days", "Status"]
for ci, h in enumerate(headers_fixed, 1):
    set_cell(ws, 1, ci, h, bold=True, color=WHITE, bg=HEADER_FILL, align="center", size=11)

# ── Day header columns (cols 6 … 6+total-1) ────
DAY_COL_W = 2.5
DAY_OFFSET = 5   # status col is 5, days start at col 6
for d in range(1, total + 1):
    col = DAY_OFFSET + d
    ltr = get_column_letter(col)
    ws.column_dimensions[ltr].width = DAY_COL_W
    set_cell(ws, 1, col, d, bold=True, color=WHITE, bg=HEADER_FILL, align="center", size=9)

# ── Task rows ──────────────────────────────────
for row_idx, tid in enumerate(range(1, num_tasks + 1), start=2):
    t = task_map[tid]
    status  = t[5]
    is_done = (status == "DONE")
    bg = ALT_ROW if row_idx % 2 == 0 else WHITE
    ws.row_dimensions[row_idx].height = 16

    set_cell(ws, row_idx, 1, tid,  align="center", bg=bg, size=9)
    # Strike-through task name if done
    name_cell = ws.cell(row=row_idx, column=2, value=t[1])
    name_cell.font = Font(
        size=9,
        strike=is_done,
        color="888888" if is_done else "000000"
    )
    name_cell.alignment = Alignment(horizontal="left", vertical="center")
    name_cell.fill = PatternFill("solid", fgColor=bg)
    name_cell.border = border

    phase_color = PHASE_COLORS.get(t[2], "888888")
    set_cell(ws, row_idx, 3, t[2], color=WHITE, bg=phase_color, align="center", size=9)
    set_cell(ws, row_idx, 4, t[3], align="center", bg=bg, size=9)

    # Status cell
    s_label, s_fcolor, s_bg = STATUS_COLORS.get(status, ("", "000000", WHITE))
    set_cell(ws, row_idx, 5, s_label, bold=(status == "DONE"), color=s_fcolor,
             bg=s_bg, align="center", size=9)

    # Bar color: green if done, red/blue if critical/normal pending
    if is_done:
        bar_color = GANTT_DONE_CRIT if t[11] else GANTT_DONE
    else:
        bar_color = GANTT_CRIT if t[11] else GANTT_BAR

    for d in range(t[6] + 1, t[7] + 1):
        col = DAY_OFFSET + d
        cell_fill(ws, row_idx, col, bar_color)
        ws.cell(row=row_idx, column=col).border = border

    for d in range(1, total + 1):
        if not (t[6] < d <= t[7]):
            c = ws.cell(row=row_idx, column=DAY_OFFSET + d)
            if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                c.fill = PatternFill("solid", fgColor=bg)
            c.border = border

# ── Progress summary row ───────────────────────
done_count = sum(1 for t in tasks if t[5] == "DONE")
inprog_count = sum(1 for t in tasks if t[5] == "IN PROGRESS")
pend_count = sum(1 for t in tasks if t[5] == "PENDING")
prog_row = num_tasks + 3

set_cell(ws, prog_row, 1,
         f"Progress: {done_count}/{num_tasks} tasks complete"
         f"  |  ✔ Done: {done_count}  ·  ⏳ In Progress: {inprog_count}  ·  ○ Pending: {pend_count}",
         bold=True, color=WHITE, bg="1E4620", align="left", size=10)
ws.merge_cells(f"A{prog_row}:{get_column_letter(DAY_OFFSET + total)}{prog_row}")
ws.row_dimensions[prog_row].height = 20

# ── Legend ─────────────────────────────────────
leg_row = prog_row + 2
set_cell(ws, leg_row, 1, "Legend:", bold=True, size=10)
set_cell(ws, leg_row, 2, "✔ Completed Task",        color=WHITE, bg=GANTT_DONE,      size=9, align="center")
set_cell(ws, leg_row, 3, "✔ Completed (Critical)",  color=WHITE, bg=GANTT_DONE_CRIT, size=9, align="center")
set_cell(ws, leg_row, 4, "Critical Path",            color=WHITE, bg=GANTT_CRIT,      size=9, align="center")
set_cell(ws, leg_row, 5, "Normal Task",              color=WHITE, bg=GANTT_BAR,       size=9, align="center")

wb_g.save(r"C:\Users\Hyunwoo Jang\Documents\Anthropic\DS440 Assignment\MHIDSS_Gantt_Chart.xlsx")
print("Gantt saved")


# ══════════════════════════════════════════════
#  PERT CHART (table form)
# ══════════════════════════════════════════════
wb_p = openpyxl.Workbook()
ws2 = wb_p.active
ws2.title = "PERT Chart"

ws2.row_dimensions[1].height = 28

col_widths  = [5, 38, 14, 9, 18, 9, 9, 9, 9, 9, 14, 14]
pert_headers = ["#", "Task Name", "Phase", "Duration\n(days)",
                "Predecessors", "ES", "EF", "LS", "LF", "Slack", "Critical?", "Status"]

for ci, (w, h) in enumerate(zip(col_widths, pert_headers), 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
    set_cell(ws2, 1, ci, h, bold=True, color=WHITE, bg=HEADER_FILL,
             align="center", wrap=True, size=10)

for row_idx, tid in enumerate(range(1, num_tasks + 1), start=2):
    t = task_map[tid]
    ws2.row_dimensions[row_idx].height = 18
    status  = t[5]
    is_done = (status == "DONE")
    crit    = t[11]

    # Row background: done=green tint, critical=red tint, else alt
    bg = ALT_ROW if row_idx % 2 == 0 else WHITE
    if is_done:
        row_bg = "E8F5E9"   # light green
    elif crit:
        row_bg = "FFE0E0"   # light red
    else:
        row_bg = bg

    preds_str = ", ".join(str(p) for p in t[4]) if t[4] else "—"
    vals = [tid, t[1], t[2], t[3], preds_str,
            t[6], t[7], t[8], t[9], t[10],
            "YES ★" if crit else "no"]

    phase_color = PHASE_COLORS.get(t[2], "888888")

    for ci, v in enumerate(vals, 1):
        if ci == 2:  # task name — strike if done
            c = ws2.cell(row=row_idx, column=ci, value=v)
            c.font = Font(size=9, strike=is_done, color="888888" if is_done else "000000")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.border = border
        elif ci == 3:   # phase
            set_cell(ws2, row_idx, ci, v, color=WHITE, bg=phase_color, align="center", size=9)
        elif ci == 11 and crit:
            set_cell(ws2, row_idx, ci, v, bold=True, color=WHITE, bg="C00000", align="center", size=9)
        else:
            align = "left" if ci in (5,) else "center"
            set_cell(ws2, row_idx, ci, v, bg=row_bg, align=align, size=9)

    # Status column (col 12)
    s_label, s_fcolor, s_bg = STATUS_COLORS.get(status, ("", "000000", WHITE))
    set_cell(ws2, row_idx, 12, s_label, bold=(status == "DONE"),
             color=s_fcolor, bg=s_bg, align="center", size=9)

# ── Progress summary ───────────────────────────
done_count   = sum(1 for t in tasks if t[5] == "DONE")
inprog_count = sum(1 for t in tasks if t[5] == "IN PROGRESS")
pend_count   = sum(1 for t in tasks if t[5] == "PENDING")
prog_row2 = num_tasks + 3

ws2.merge_cells(f"A{prog_row2}:L{prog_row2}")
set_cell(ws2, prog_row2, 1,
         f"Progress: {done_count}/{num_tasks} tasks complete  "
         f"|  ✔ Done: {done_count}  ·  ⏳ In Progress: {inprog_count}  ·  ○ Pending: {pend_count}",
         bold=True, color=WHITE, bg="1E4620", align="left", size=10)
ws2.row_dimensions[prog_row2].height = 20

# ── Critical Path summary ──────────────────────
cp_row = prog_row2 + 2
cp_tasks = [tid for tid in range(1, num_tasks + 1) if task_map[tid][11]]
cp_str = " → ".join(str(t) for t in cp_tasks)
ws2.merge_cells(f"A{cp_row}:L{cp_row}")
set_cell(ws2, cp_row, 1,
         f"Critical Path ({total} days):  {cp_str}",
         bold=True, color=WHITE, bg="C00000", align="left", size=10)
ws2.row_dimensions[cp_row].height = 20

# ── PERT legend ────────────────────────────────
leg2 = cp_row + 2
set_cell(ws2, leg2,   1, "Node format:", bold=True, size=9)
set_cell(ws2, leg2,   2,
         "ES = Early Start  |  EF = Early Finish  |  LS = Late Start  |  LF = Late Finish  |  Slack = LS − ES",
         size=9)
set_cell(ws2, leg2+1, 1, "Row colors:", bold=True, size=9)
set_cell(ws2, leg2+1, 2,
         "Green tint = Completed  |  Red tint = Critical Path (pending)  |  White/Gray = Normal",
         size=9)
set_cell(ws2, leg2+2, 1, "Total duration:", bold=True, size=9)
set_cell(ws2, leg2+2, 2, f"{total} days", bold=True, color="C00000", size=9)

wb_p.save(r"C:\Users\Hyunwoo Jang\Documents\Anthropic\DS440 Assignment\MHIDSS_PERT_Chart.xlsx")
print("PERT saved")
